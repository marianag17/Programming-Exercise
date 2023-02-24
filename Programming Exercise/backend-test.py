import os
import shutil
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import openpyxl
from openpyxl import load_workbook
from pathlib import Path

# Specify the folder to move processed files to
PROCESSED_FOLDER = Path(__file__) / "../processed/"

# Specify the folder to move non-applicable files to
NOT_APPLICABLE_FOLDER = Path(__file__) / "../not_applicable/"

# Specify the name of the master workbook file
MASTER_WORKBOOK_NAME = 'master_workbook.xlsx'

# Define a class to handle file system events
class FileHandler(FileSystemEventHandler):
    def on_created(self, event):
        try:
            
            # Check if the event is a file and not a folder
            if not event.is_directory:
                file_path = event.src_path
                file_name, file_extension = os.path.splitext(file_path)

                # Check if the file is an Excel file
                if file_extension.lower() in ('.xls', '.xlsx', '.xlsm'):
                    # Load the file
                    workbook = load_workbook(file_path)

                    # see every sheet and copy it to the master workbook
                    master_workbook = load_workbook(MASTER_WORKBOOK_NAME)
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        master_worksheet = master_workbook.create_sheet(title=sheet_name)
                        for row in worksheet.iter_rows(values_only=True):
                            master_worksheet.append(row)

                    # Save the master workbook
                    master_workbook.save(MASTER_WORKBOOK_NAME)

                    # Move the file to the processed folder
                    shutil.move(file_path, PROCESSED_FOLDER)
                    print('File moved to processed folder')
                else:
                    # Move the file to the not_applicable folder
                    shutil.move(file_path, NOT_APPLICABLE_FOLDER)
                    print('File moved to not applicable folder')
        except:
            print("Please verify that you have permissions for the files or that you entered a valid path.")
            
if __name__ == "__main__":
    # Create the master workbook if it doesn't exist
    if not os.path.exists(MASTER_WORKBOOK_NAME):
        master_workbook = openpyxl.Workbook()
        master_workbook.save(MASTER_WORKBOOK_NAME)

    print('''
______    _     _                           _       _               
|  ___|  | |   | |                         | |     | |              
| |_ ___ | | __| | ___ _ __  __      ____ _| |_ ___| |__   ___ _ __ 
|  _/ _ \| |/ _` |/ _ \ '__| \ \ /\ / / _` | __/ __| '_ \ / _ \ '__|
| || (_) | | (_| |  __/ |     \ V  V / (_| | || (__| | | |  __/ |   
\_| \___/|_|\__,_|\___|_|      \_/\_/ \__,_|\__\___|_| |_|\___|_|   
                                                                    
                                                                    ''')
    # Specify the folder to watch
    WATCH_FOLDER = input("Please, enter the path of the foder to watch: \n")
    try:
        # Set up the event handler and observer
        event_handler = FileHandler()
        observer = Observer()
        observer.schedule(event_handler, WATCH_FOLDER, recursive=True)

        # Start the observer
        observer.start()

        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            observer.stop()

        observer.join()
    except:
        print("Please verify that you have permissions for the files or that you entered a valid path.")
        
    
    
